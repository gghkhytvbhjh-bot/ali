import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

from kivy.config import Config
Config.set('graphics', 'resizable', True)

from kivymd.app import MDApp
from kivymd.uix.screen import MDScreen
from kivymd.uix.screenmanager import MDScreenManager
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.toolbar import MDTopAppBar
from kivymd.uix.button import MDRaisedButton, MDFlatButton
from kivymd.uix.textfield import MDTextField
from kivymd.uix.dialog import MDDialog
from kivymd.uix.label import MDLabel
from kivymd.uix.card import MDCard
from kivymd.uix.menu import MDDropdownMenu
from kivy.uix.scrollview import ScrollView
from kivy.metrics import dp


def get_data_path():
    if 'ANDROID_ARGUMENT' in os.environ:
        from android.storage import app_storage_path
        return app_storage_path()
    return os.path.dirname(os.path.abspath(__file__))


RELATIONS = [
    "Lui-meme", "Elle-meme",
    "Pere", "Mere", "Frere", "Soeur",
    "Epoux", "Epouse",
    "Fils", "Fille",
    "Cousin", "Cousine",
    "Ami", "Amie",
    "Collegue", "Voisin", "Autre"
]


class AccueilScreen(MDScreen):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.name = "accueil"
        self.app_ref = app

        layout = MDBoxLayout(orientation='vertical')
        toolbar = MDTopAppBar(title="AidAdha - Controle Distribution")
        toolbar.elevation = 4
        layout.add_widget(toolbar)

        content = MDBoxLayout(
            orientation='vertical',
            padding=dp(24),
            spacing=dp(16)
        )

        self.label_statut = MDLabel(
            text="Aucun fichier beneficiaires charge",
            halign="center",
            theme_text_color="Secondary",
            size_hint_y=None,
            height=dp(40)
        )
        content.add_widget(self.label_statut)

        btn_import = MDRaisedButton(
            text="Importer la liste Excel des beneficiaires",
            on_release=self.importer_liste,
            md_bg_color=(0.13, 0.59, 0.95, 1)
        )
        content.add_widget(btn_import)

        btn_controle = MDRaisedButton(
            text="Commencer le Controle",
            on_release=self.aller_controle,
            md_bg_color=(0.13, 0.77, 0.37, 1)
        )
        content.add_widget(btn_controle)

        btn_rapport = MDRaisedButton(
            text="Voir les Rapports",
            on_release=self.aller_rapport,
            md_bg_color=(0.96, 0.49, 0.0, 1)
        )
        content.add_widget(btn_rapport)

        self.label_stats = MDLabel(
            text="",
            halign="center",
            theme_text_color="Primary",
            size_hint_y=None,
            height=dp(50)
        )
        content.add_widget(self.label_stats)

        layout.add_widget(content)
        self.add_widget(layout)
        self.mettre_a_jour_stats()

    def importer_liste(self, instance):
        self.dialog_import = MDDialog(
            title="Importer la liste",
            text="Placez votre fichier Excel nomme 'beneficiaires.xlsx' dans le meme dossier que l'application, puis confirmez.",
            buttons=[
                MDFlatButton(text="Annuler", on_release=lambda x: self.dialog_import.dismiss()),
                MDRaisedButton(text="Confirmer", on_release=self.confirmer_import)
            ]
        )
        self.dialog_import.open()

    def confirmer_import(self, instance):
        self.dialog_import.dismiss()
        chemin = os.path.join(get_data_path(), "beneficiaires.xlsx")
        if os.path.exists(chemin):
            try:
                wb = load_workbook(chemin)
                ws = wb.active
                total = ws.max_row - 1
                self.label_statut.text = f"Liste chargee : {total} beneficiaires"
                self.label_statut.theme_text_color = "Custom"
                self.label_statut.text_color = (0.13, 0.77, 0.37, 1)
                self.mettre_a_jour_stats()
                self.afficher_dialogue("Succes", f"{total} beneficiaires importes !")
            except Exception as e:
                self.afficher_dialogue("Erreur", str(e))
        else:
            self.afficher_dialogue("Introuvable", f"Placez 'beneficiaires.xlsx' dans :\n{get_data_path()}")

    def mettre_a_jour_stats(self):
        retrait_path = os.path.join(get_data_path(), "retraits.xlsx")
        if os.path.exists(retrait_path):
            try:
                wb = load_workbook(retrait_path)
                ws = wb.active
                retraits = list(ws.iter_rows(min_row=2, values_only=True))
                retraits = [r for r in retraits if r and any(r)]
                total = len(retraits)
                aujourd_hui = datetime.now().strftime("%Y-%m-%d")
                jour = sum(1 for r in retraits if str(r[0]).startswith(aujourd_hui))
                self.label_stats.text = f"Aujourd'hui : {jour} retrait(s)\nTotal mission : {total} retrait(s)"
            except Exception:
                pass

    def aller_controle(self, instance):
        self.app_ref.root.current = "controle"

    def aller_rapport(self, instance):
        self.app_ref.root.get_screen("rapport").charger_rapport()
        self.app_ref.root.current = "rapport"

    def afficher_dialogue(self, titre, message):
        d = MDDialog(title=titre, text=message,
                     buttons=[MDFlatButton(text="OK", on_release=lambda x: d.dismiss())])
        d.open()


class ControleScreen(MDScreen):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.name = "controle"
        self.app_ref = app
        self.beneficiaire_trouve = None
        self.relation_selectionnee = "Lui-meme"

        layout = MDBoxLayout(orientation='vertical')
        toolbar = MDTopAppBar(
            title="Controle du Jeton",
            left_action_items=[["arrow-left", lambda x: self.retour()]]
        )
        toolbar.elevation = 4
        layout.add_widget(toolbar)

        scroll = ScrollView()
        content = MDBoxLayout(
            orientation='vertical',
            padding=dp(16),
            spacing=dp(12),
            size_hint_y=None
        )
        content.bind(minimum_height=content.setter('height'))

        # Recherche
        content.add_widget(MDLabel(
            text="Rechercher par Nom, CIN ou Numero de serie :",
            theme_text_color="Secondary",
            size_hint_y=None, height=dp(30)
        ))

        self.input_recherche = MDTextField(
            hint_text="Tapez le nom, CIN ou numero de serie...",
            size_hint_y=None, height=dp(50)
        )
        content.add_widget(self.input_recherche)

        content.add_widget(MDRaisedButton(
            text="Rechercher",
            on_release=self.rechercher,
            md_bg_color=(0.13, 0.59, 0.95, 1),
            size_hint_y=None, height=dp(46)
        ))

        # Carte resultat beneficiaire
        self.carte_resultat = MDCard(
            orientation='vertical',
            padding=dp(14),
            spacing=dp(8),
            size_hint_y=None,
            height=dp(80),
            elevation=3
        )
        self.carte_resultat.opacity = 0

        self.lbl_nom = MDLabel(text="", theme_text_color="Primary", font_style="H6",
                               size_hint_y=None, height=dp(30))
        self.lbl_cin = MDLabel(text="", theme_text_color="Secondary",
                               size_hint_y=None, height=dp(25))
        self.lbl_serie = MDLabel(text="", theme_text_color="Secondary",
                                 size_hint_y=None, height=dp(25))
        self.lbl_statut = MDLabel(text="", size_hint_y=None, height=dp(25))

        self.carte_resultat.add_widget(self.lbl_nom)
        self.carte_resultat.add_widget(self.lbl_cin)
        self.carte_resultat.add_widget(self.lbl_serie)
        self.carte_resultat.add_widget(self.lbl_statut)
        content.add_widget(self.carte_resultat)

        # Section retrait par un tiers
        self.carte_tiers = MDCard(
            orientation='vertical',
            padding=dp(14),
            spacing=dp(10),
            size_hint_y=None,
            height=dp(280),
            elevation=3
        )
        self.carte_tiers.opacity = 0

        self.carte_tiers.add_widget(MDLabel(
            text="Retire par :",
            theme_text_color="Primary",
            font_style="Subtitle1",
            size_hint_y=None, height=dp(30)
        ))

        # Bouton selection relation
        self.btn_relation = MDRaisedButton(
            text="Relation : Lui-meme",
            on_release=self.ouvrir_menu_relation,
            md_bg_color=(0.4, 0.4, 0.8, 1),
            size_hint_y=None, height=dp(44)
        )
        self.carte_tiers.add_widget(self.btn_relation)

        # Champs tiers (masques par defaut)
        self.input_nom_tiers = MDTextField(
            hint_text="Nom et Prenom du representant",
            size_hint_y=None, height=dp(50)
        )
        self.input_cin_tiers = MDTextField(
            hint_text="CIN du representant",
            size_hint_y=None, height=dp(50)
        )
        self.carte_tiers.add_widget(self.input_nom_tiers)
        self.carte_tiers.add_widget(self.input_cin_tiers)

        # Bouton confirmer retrait
        self.btn_valider = MDRaisedButton(
            text="Confirmer le retrait du mouton",
            on_release=self.confirmer_retrait,
            md_bg_color=(0.13, 0.77, 0.37, 1),
            size_hint_y=None, height=dp(46)
        )
        self.carte_tiers.add_widget(self.btn_valider)
        content.add_widget(self.carte_tiers)

        scroll.add_widget(content)
        layout.add_widget(scroll)
        self.add_widget(layout)

        # Menu dropdown relations
        menu_items = [
            {"text": r, "on_release": lambda x=r: self.selectionner_relation(x)}
            for r in RELATIONS
        ]
        self.menu_relation = MDDropdownMenu(
            caller=self.btn_relation,
            items=menu_items,
            width_mult=4
        )

    def ouvrir_menu_relation(self, instance):
        self.menu_relation.open()

    def selectionner_relation(self, relation):
        self.relation_selectionnee = relation
        self.btn_relation.text = f"Relation : {relation}"
        self.menu_relation.dismiss()

        # Afficher/masquer champs tiers
        est_tiers = relation not in ["Lui-meme", "Elle-meme"]
        self.input_nom_tiers.opacity = 1 if est_tiers else 0
        self.input_cin_tiers.opacity = 1 if est_tiers else 0
        if not est_tiers:
            self.input_nom_tiers.text = ""
            self.input_cin_tiers.text = ""

    def retour(self):
        self.app_ref.root.get_screen("accueil").mettre_a_jour_stats()
        self.app_ref.root.current = "accueil"

    def rechercher(self, instance):
        terme = self.input_recherche.text.strip().lower()
        if not terme:
            self.afficher_dialogue("Attention", "Veuillez saisir un nom, CIN ou numero de serie.")
            return

        chemin = os.path.join(get_data_path(), "beneficiaires.xlsx")
        if not os.path.exists(chemin):
            self.afficher_dialogue("Erreur", "Liste introuvable. Importez d'abord la liste depuis l'accueil.")
            return

        try:
            wb = load_workbook(chemin)
            ws = wb.active
            self.beneficiaire_trouve = None

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                nom = str(row[0]).lower() if row[0] else ""
                prenom = str(row[1]).lower() if row[1] else ""
                cin = str(row[2]).lower() if row[2] else ""
                serie = str(row[3]).lower() if row[3] else ""

                if (terme in nom or terme in prenom or
                        terme in (nom + " " + prenom) or
                        terme == cin or terme == serie):
                    self.beneficiaire_trouve = {
                        "nom": str(row[0]) if row[0] else "",
                        "prenom": str(row[1]) if row[1] else "",
                        "cin": str(row[2]) if row[2] else "",
                        "serie": str(row[3]) if row[3] else ""
                    }
                    break

            if self.beneficiaire_trouve:
                self.afficher_resultat()
            else:
                self.carte_resultat.opacity = 0
                self.carte_tiers.opacity = 0
                self.afficher_dialogue("Introuvable", f"Aucun beneficiaire trouve pour : {terme}")

        except Exception as e:
            self.afficher_dialogue("Erreur", str(e))

    def afficher_resultat(self):
        b = self.beneficiaire_trouve
        self.lbl_nom.text = f"{b['nom']} {b['prenom']}"
        self.lbl_cin.text = f"CIN : {b['cin']}"
        self.lbl_serie.text = f"N Serie : {b['serie']}"

        deja_retire = self.verifier_retrait(b['serie'])
        if deja_retire:
            self.lbl_statut.text = f"DEJA RETIRE le {deja_retire}"
            self.lbl_statut.theme_text_color = "Custom"
            self.lbl_statut.text_color = (0.9, 0.1, 0.1, 1)
            self.carte_tiers.opacity = 0
        else:
            self.lbl_statut.text = "En attente de retrait"
            self.lbl_statut.theme_text_color = "Custom"
            self.lbl_statut.text_color = (0.13, 0.77, 0.37, 1)
            self.carte_tiers.opacity = 1
            # Reset relation
            self.relation_selectionnee = "Lui-meme"
            self.btn_relation.text = "Relation : Lui-meme"
            self.input_nom_tiers.opacity = 0
            self.input_cin_tiers.opacity = 0
            self.input_nom_tiers.text = ""
            self.input_cin_tiers.text = ""

        self.carte_resultat.opacity = 1

    def verifier_retrait(self, serie):
        retrait_path = os.path.join(get_data_path(), "retraits.xlsx")
        if not os.path.exists(retrait_path):
            return None
        try:
            wb = load_workbook(retrait_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and str(row[2]) == str(serie):
                    return str(row[0])
        except Exception:
            pass
        return None

    def confirmer_retrait(self, instance):
        if not self.beneficiaire_trouve:
            return
        b = self.beneficiaire_trouve
        est_tiers = self.relation_selectionnee not in ["Lui-meme", "Elle-meme"]

        if est_tiers:
            nom_tiers = self.input_nom_tiers.text.strip()
            cin_tiers = self.input_cin_tiers.text.strip()
            if not nom_tiers or not cin_tiers:
                self.afficher_dialogue("Attention", "Veuillez saisir le nom et CIN du representant.")
                return
            msg = (f"Confirmer que {b['nom']} {b['prenom']} (Serie: {b['serie']})\n"
                   f"est represente par : {nom_tiers}\n"
                   f"CIN representant : {cin_tiers}\n"
                   f"Relation : {self.relation_selectionnee}")
        else:
            msg = (f"Confirmer que {b['nom']} {b['prenom']}\n"
                   f"(Serie: {b['serie']}) retire son mouton en personne.")

        self.dialog_confirm = MDDialog(
            title="Confirmer le retrait",
            text=msg,
            buttons=[
                MDFlatButton(text="Annuler", on_release=lambda x: self.dialog_confirm.dismiss()),
                MDRaisedButton(text="OUI CONFIRMER", on_release=self.enregistrer_retrait)
            ]
        )
        self.dialog_confirm.open()

    def enregistrer_retrait(self, instance):
        self.dialog_confirm.dismiss()
        b = self.beneficiaire_trouve
        est_tiers = self.relation_selectionnee not in ["Lui-meme", "Elle-meme"]
        nom_tiers = self.input_nom_tiers.text.strip() if est_tiers else ""
        cin_tiers = self.input_cin_tiers.text.strip() if est_tiers else ""

        retrait_path = os.path.join(get_data_path(), "retraits.xlsx")
        try:
            if os.path.exists(retrait_path):
                wb = load_workbook(retrait_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Retraits"
                ws.append([
                    "Date/Heure", "Nom Beneficiaire", "N Serie", "CIN Beneficiaire",
                    "Retire par", "Nom Representant", "CIN Representant", "Relation"
                ])

            ws.append([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                f"{b['nom']} {b['prenom']}",
                b['serie'],
                b['cin'],
                self.relation_selectionnee,
                nom_tiers if est_tiers else f"{b['nom']} {b['prenom']}",
                cin_tiers if est_tiers else b['cin'],
                self.relation_selectionnee
            ])
            wb.save(retrait_path)

            self.lbl_statut.text = f"RETIRE le {datetime.now().strftime('%H:%M')}"
            self.lbl_statut.text_color = (0.9, 0.1, 0.1, 1)
            self.carte_tiers.opacity = 0
            self.input_recherche.text = ""
            self.beneficiaire_trouve = None

            if est_tiers:
                self.afficher_dialogue("Succes !",
                    f"Retrait confirme pour {b['nom']} {b['prenom']}\n"
                    f"Represente par : {nom_tiers}\n"
                    f"Relation : {self.relation_selectionnee}")
            else:
                self.afficher_dialogue("Succes !",
                    f"Retrait confirme pour\n{b['nom']} {b['prenom']}")

        except Exception as e:
            self.afficher_dialogue("Erreur", str(e))

    def afficher_dialogue(self, titre, message):
        d = MDDialog(title=titre, text=message,
                     buttons=[MDFlatButton(text="OK", on_release=lambda x: d.dismiss())])
        d.open()


class RapportScreen(MDScreen):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.name = "rapport"
        self.app_ref = app

        layout = MDBoxLayout(orientation='vertical')
        toolbar = MDTopAppBar(
            title="Rapports de Distribution",
            left_action_items=[["arrow-left", lambda x: setattr(self.app_ref.root, 'current', 'accueil')]]
        )
        toolbar.elevation = 4
        layout.add_widget(toolbar)

        scroll = ScrollView()
        self.content = MDBoxLayout(
            orientation='vertical',
            padding=dp(16),
            spacing=dp(10),
            size_hint_y=None
        )
        self.content.bind(minimum_height=self.content.setter('height'))
        scroll.add_widget(self.content)
        layout.add_widget(scroll)
        self.add_widget(layout)

    def charger_rapport(self):
        self.content.clear_widgets()
        retrait_path = os.path.join(get_data_path(), "retraits.xlsx")

        if not os.path.exists(retrait_path):
            self.content.add_widget(MDLabel(
                text="Aucun retrait enregistre.",
                halign="center", theme_text_color="Secondary"
            ))
            return

        try:
            wb = load_workbook(retrait_path)
            ws = wb.active
            retraits = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and any(r)]
            total = len(retraits)
            aujourd_hui = datetime.now().strftime("%Y-%m-%d")
            jour_count = sum(1 for r in retraits if str(r[0]).startswith(aujourd_hui))

            # Compte par relation
            relations_count = {}
            tiers_count = 0
            for r in retraits:
                rel = str(r[4]) if r[4] else "Inconnu"
                relations_count[rel] = relations_count.get(rel, 0) + 1
                if rel not in ["Lui-meme", "Elle-meme"]:
                    tiers_count += 1

            # Par jour
            jours = {}
            for r in retraits:
                jour = str(r[0])[:10] if r[0] else "?"
                jours[jour] = jours.get(jour, 0) + 1

            def ajouter(texte, couleur=None, taille="Body1"):
                lbl = MDLabel(
                    text=texte,
                    size_hint_y=None,
                    height=dp(30),
                    font_style=taille
                )
                if couleur:
                    lbl.theme_text_color = "Custom"
                    lbl.text_color = couleur
                else:
                    lbl.theme_text_color = "Secondary"
                self.content.add_widget(lbl)

            ajouter(f"RAPPORT DU {aujourd_hui}", (0.13, 0.59, 0.95, 1), "H6")
            ajouter(f"Retraits aujourd'hui : {jour_count}", (0.13, 0.77, 0.37, 1))
            ajouter(f"Dont retires par un tiers : {tiers_count}", (0.96, 0.49, 0.0, 1))
            ajouter("─" * 40)

            ajouter("PAR JOUR :", (0.2, 0.2, 0.2, 1), "Subtitle1")
            for jour, count in sorted(jours.items()):
                ajouter(f"  {jour} : {count} retrait(s)")

            ajouter("─" * 40)
            ajouter("PAR RELATION :", (0.2, 0.2, 0.2, 1), "Subtitle1")
            for rel, count in sorted(relations_count.items(), key=lambda x: -x[1]):
                ajouter(f"  {rel} : {count}")

            ajouter("─" * 40)
            ajouter(f"TOTAL MISSION : {total} retrait(s)", (0.96, 0.49, 0.0, 1), "H6")
            ajouter("─" * 40)

            ajouter("DERNIERS RETRAITS :", (0.2, 0.2, 0.2, 1), "Subtitle1")
            for r in reversed(retraits[-20:]):
                rel = str(r[4]) if r[4] else ""
                rep = f" | par {r[5]} ({rel})" if rel not in ["Lui-meme", "Elle-meme"] else ""
                ajouter(f"  {str(r[0])[5:16]} | {r[1]} | Serie:{r[2]}{rep}", None, "Caption")

        except Exception as e:
            self.content.add_widget(MDLabel(text=f"Erreur : {str(e)}"))


class AidAdhaApp(MDApp):
    def build(self):
        self.theme_cls.primary_palette = "Blue"
        self.theme_cls.theme_style = "Light"
        sm = MDScreenManager()
        sm.add_widget(AccueilScreen(app=self))
        sm.add_widget(ControleScreen(app=self))
        sm.add_widget(RapportScreen(app=self))
        sm.current = "accueil"
        return sm


if __name__ == "__main__":
    AidAdhaApp().run()
